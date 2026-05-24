from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment


input_file = "原始成绩.xlsx"
output_file = "统计结果.xlsx"


# 读取原始 Excel
input_wb = load_workbook(input_file)
input_ws = input_wb.active


# 创建结果 Excel
output_wb = Workbook()

# 第一个 sheet：明细表
detail_ws = output_wb.active
detail_ws.title = "成绩明细"

# 第二个 sheet：汇总表
summary_ws = output_wb.create_sheet("汇总统计")


# 写明细表表头
detail_headers = ["姓名", "单选得分", "多选得分", "总分", "是否及格"]
detail_ws.append(detail_headers)


total_scores = []
passed_count = 0
failed_count = 0
student_count = 0


# 逐行读取数据
for row in input_ws.iter_rows(min_row=2, values_only=True):
    name = row[0]
    single_score = row[1]
    multiple_score = row[2]

    # 跳过完全空行
    if name is None and single_score is None and multiple_score is None:
        continue

    # 如果姓名为空，跳过这一行
    if name is None:
        continue

    # 如果分数为空，按 0 分处理
    if single_score is None:
        single_score = 0

    if multiple_score is None:
        multiple_score = 0

    total_score = single_score + multiple_score

    if total_score >= 60:
        passed = "及格"
        passed_count += 1
    else:
        passed = "不及格"
        failed_count += 1

    student_count += 1
    total_scores.append(total_score)

    detail_ws.append([
        name,
        single_score,
        multiple_score,
        total_score,
        passed
    ])


# 计算汇总数据
if student_count > 0:
    average_score = sum(total_scores) / student_count
    max_score = max(total_scores)
    min_score = min(total_scores)
else:
    average_score = 0
    max_score = 0
    min_score = 0


# 写汇总表
summary_ws.append(["统计项目", "数值"])
summary_ws.append(["总人数", student_count])
summary_ws.append(["及格人数", passed_count])
summary_ws.append(["不及格人数", failed_count])
summary_ws.append(["平均分", round(average_score, 2)])
summary_ws.append(["最高分", max_score])
summary_ws.append(["最低分", min_score])


# 设置样式函数
def format_sheet(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = 14


format_sheet(detail_ws)
format_sheet(summary_ws)


# 保存结果
output_wb.save(output_file)

print("统计完成，已生成：", output_file)
